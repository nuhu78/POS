import io
from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from rest_framework.viewsets import ModelViewSet
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import filters, status
from django.db.models import F
from .models import Product
from .serializers import ProductSerializer
from .permissions import AdminOrReadOnly
from apps.categories.models import Category


EXPORT_HEADERS = ["SKU", "Name", "Category", "Purchase Price", "Selling Price", "Stock", "Low Stock Threshold", "Status"]
CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class ProductViewSet(ModelViewSet):
    queryset = Product.objects.all().order_by("name")
    serializer_class = ProductSerializer
    permission_classes = [AdminOrReadOnly]
    search_fields = ["name", "sku"]
    filterset_fields = ["category", "status"]

    @action(detail=False, methods=["get"])
    def low_stock(self, request):
        products = Product.objects.filter(
            stock__lte=F("low_stock_threshold"), status="active"
        ).order_by("stock")
        return Response(ProductSerializer(products, many=True).data)

    @action(detail=False, methods=["get"])
    def export(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(EXPORT_HEADERS)

        products = Product.objects.select_related("category").all().order_by("name")
        for p in products:
            ws.append([
                p.sku,
                p.name,
                p.category.name,
                str(p.purchase_price),
                str(p.selling_price),
                p.stock,
                p.low_stock_threshold,
                p.status,
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type=CONTENT_TYPE,
        )
        response["Content-Disposition"] = 'attachment; filename="products.xlsx"'
        return response

    @action(detail=False, methods=["post"], url_path="import")
    def import_products(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response(
                {"error": {"code": "INVALID_FILE", "message": "No file provided."}},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not file.name.endswith(".xlsx"):
            return Response(
                {"error": {"code": "INVALID_FILE", "message": "Only .xlsx files are supported."}},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if file.content_type != CONTENT_TYPE:
            return Response(
                {"error": {"code": "INVALID_FILE", "message": "Invalid file content type."}},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = load_workbook(file, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
        except Exception:
            return Response(
                {"error": {"code": "INVALID_FILE", "message": "Could not read the file. It may be corrupted."}},
                status=status.HTTP_400_BAD_REQUEST,
            )

        processed = 0
        added = 0
        updated = 0
        skipped = []

        category_cache = {}

        def get_category(name):
            if not name or not str(name).strip():
                return None
            name = str(name).strip()
            if name not in category_cache:
                try:
                    category_cache[name] = Category.objects.get(name__iexact=name)
                except Category.DoesNotExist:
                    category_cache[name] = None
            return category_cache[name]

        with transaction.atomic():
            for i, row in enumerate(rows):
                row_num = i + 2
                if not row or all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
                    skipped.append({"row": row_num, "reason": "Empty row"})
                    continue

                sku = str(row[0]).strip().upper() if row[0] else ""
                if not sku:
                    skipped.append({"row": row_num, "reason": "Missing SKU"})
                    continue

                name = str(row[1]).strip() if row[1] else ""
                if not name:
                    skipped.append({"row": row_num, "reason": "Missing name"})
                    continue

                category_name = str(row[2]).strip() if row[2] else ""
                category = get_category(category_name) if category_name else None
                if category_name and not category:
                    skipped.append({"row": row_num, "reason": f"Category '{category_name}' not found"})
                    continue

                try:
                    purchase_price = float(row[3]) if row[3] is not None else 0
                    selling_price = float(row[4]) if row[4] is not None else 0
                    stock = int(float(row[5])) if row[5] is not None else 0
                    threshold = int(float(row[6])) if row[6] is not None else 5
                except (ValueError, TypeError):
                    skipped.append({"row": row_num, "reason": "Invalid numeric value"})
                    continue

                status_val = str(row[7]).strip().lower() if row[7] else "active"
                if status_val not in ("active", "inactive"):
                    status_val = "active"

                data = {
                    "name": name,
                    "category": category,
                    "purchase_price": purchase_price,
                    "selling_price": selling_price,
                    "stock": stock,
                    "low_stock_threshold": threshold,
                    "status": status_val,
                }

                existing = Product.objects.filter(sku=sku).first()
                if existing:
                    for field, value in data.items():
                        setattr(existing, field, value)
                    existing.save()
                    updated += 1
                else:
                    Product.objects.create(sku=sku, **data)
                    added += 1

                processed += 1

        return Response({
            "processed": processed,
            "added": added,
            "updated": updated,
            "skipped": skipped,
        })
